"""
모임/캘린더 라우터: 모임 CRUD, 캘린더 조회, 참석 투표.

코트 면 수(court_count)는 모임 단위로 등록되며, 이후 대진 자동생성에서
한 라운드의 동시 진행 매치 수로 사용된다.
"""
import io
from datetime import date, datetime, time

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from ..database import get_db
from ..deps import get_current_user
from ..models import AttendanceStatus, Gathering, Participant, User
from ..schemas import (
    AttendanceSummary,
    GatheringCreate,
    GatheringDetail,
    GatheringImportResult,
    GatheringImportRowError,
    GatheringRead,
    GatheringUpdate,
    ParticipantRead,
    ParticipantVote,
)

router = APIRouter(prefix="/gatherings", tags=["gatherings"])

# 엑셀 헤더(한국어) → 모임 필드명 매핑
IMPORT_HEADER_MAP = {
    "날짜": "event_date",
    "제목": "title",
    "시작시간": "start_time",
    "종료시간": "end_time",
    "장소": "location",
    "코트번호": "court_numbers",
    "최대인원": "max_participants",
    "설명": "description",
}
IMPORT_COLUMNS = list(IMPORT_HEADER_MAP.keys())
# 필드명 → 한국어 헤더 (오류 메시지 표기용)
FIELD_TO_HEADER = {v: k for k, v in IMPORT_HEADER_MAP.items()}


# --- 내부 헬퍼 --------------------------------------------------------------
def _summary(gathering: Gathering) -> AttendanceSummary:
    s = AttendanceSummary()
    for p in gathering.participants:
        if p.status == AttendanceStatus.ATTENDING:
            s.attending += 1
        elif p.status == AttendanceStatus.ABSENT:
            s.absent += 1
        else:
            s.maybe += 1
    s.total = len(gathering.participants)
    return s


def _to_read(gathering: Gathering) -> GatheringRead:
    read = GatheringRead.model_validate(gathering)
    read.attendance = _summary(gathering)
    return read


def _load_gathering(db: Session, gathering_id: int) -> Gathering:
    g = db.scalar(
        select(Gathering)
        .where(Gathering.id == gathering_id)
        .options(selectinload(Gathering.participants).selectinload(Participant.user))
    )
    if not g:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="모임을 찾을 수 없습니다.")
    return g


def _require_organizer(gathering: Gathering, user: User) -> None:
    if gathering.created_by != user.id and not user.is_admin:
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="모임 주최자 또는 관리자만 가능합니다.")


def _normalize_courts(data: dict) -> None:
    """court_numbers가 주어지면 정리하고 court_count를 그 개수로 맞춘다."""
    raw = data.get("court_numbers")
    if raw:
        labels = [s.strip() for s in str(raw).split(",") if s.strip()]
        if labels:
            data["court_numbers"] = ", ".join(labels)
            data["court_count"] = len(labels)


# 모임 시작 며칠 전부터 참석 변경(불참/미정)을 잠그는지
ATTENDANCE_LOCK_DAYS = 3


def _attendance_locked(gathering: Gathering) -> bool:
    """모임 시작 ATTENDANCE_LOCK_DAYS일 전부터(당일·지난 경우 포함) 잠금."""
    return (gathering.event_date - date.today()).days <= ATTENDANCE_LOCK_DAYS


# --- 모임 CRUD --------------------------------------------------------------
@router.post("", response_model=GatheringRead, status_code=status.HTTP_201_CREATED)
def create_gathering(
    payload: GatheringCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """모임 생성 (캘린더에 일정 등록)."""
    data = payload.model_dump()
    _normalize_courts(data)
    gathering = Gathering(**data, created_by=current_user.id)
    db.add(gathering)
    db.commit()
    return _to_read(_load_gathering(db, gathering.id))


# --- 엑셀 일괄 업로드 -------------------------------------------------------
def _cell_value(field: str, value) -> str | int:
    """엑셀 셀 값을 GatheringCreate가 받을 수 있는 형태로 정규화."""
    if field == "event_date":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip()
    if field in ("start_time", "end_time"):
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, time):
            return value.strftime("%H:%M")
        return str(value).strip()
    if field == "max_participants":
        return int(value)
    # title, location, court_numbers, description
    return str(value).strip()


@router.post("/import", response_model=GatheringImportResult)
async def import_gatherings(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """엑셀(.xlsx) 한 행 = 한 모임으로 일괄 등록. 오류 행은 건너뛰고 보고한다."""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=".xlsx 파일만 업로드할 수 있습니다.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="엑셀 파일을 읽을 수 없습니다.")

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="빈 파일입니다.")

    # 헤더명 → 열 인덱스 (알 수 없는 컬럼은 무시)
    col_to_field: dict[int, str] = {}
    for idx, name in enumerate(header):
        key = str(name).strip() if name is not None else ""
        if key in IMPORT_HEADER_MAP:
            col_to_field[idx] = IMPORT_HEADER_MAP[key]

    created = 0
    errors: list[GatheringImportRowError] = []

    for i, row in enumerate(rows, start=2):  # 2행부터 데이터
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # 완전 빈 행 skip
        try:
            data: dict = {}
            for idx, field in col_to_field.items():
                value = row[idx] if idx < len(row) else None
                if value is None or str(value).strip() == "":
                    continue
                data[field] = _cell_value(field, value)
            payload = GatheringCreate(**data)  # pydantic 검증/타입 변환 재사용
            d = payload.model_dump()
            _normalize_courts(d)
            db.add(Gathering(**d, created_by=current_user.id))
            created += 1
        except ValidationError as ve:
            parts = []
            for err in ve.errors():
                loc = err["loc"][0] if err["loc"] else ""
                label = FIELD_TO_HEADER.get(str(loc), str(loc))
                parts.append(f"{label}: {err['msg']}")
            errors.append(GatheringImportRowError(row=i, error="; ".join(parts)))
        except Exception as e:
            errors.append(GatheringImportRowError(row=i, error=str(e)))

    db.commit()
    return GatheringImportResult(created=created, failed=len(errors), errors=errors)


@router.get("/import/template")
def import_template(_: User = Depends(get_current_user)):
    """업로드용 엑셀 양식(.xlsx) 다운로드: 헤더 + 예시 1행."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "모임목록"
    ws.append(IMPORT_COLUMNS)
    ws.append(["2026-07-05", "정기 모임", "09:00", "12:00", "시민 테니스장", "3, 5", 16, "비고 예시"])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="gathering_template.xlsx"'},
    )


@router.get("", response_model=list[GatheringRead])
def list_gatherings(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
    date_from: date | None = Query(None, description="캘린더 조회 시작일"),
    date_to: date | None = Query(None, description="캘린더 조회 종료일"),
):
    """캘린더 조회: 기간으로 모임 목록을 가져온다(참석 요약 포함)."""
    stmt = select(Gathering).options(
        selectinload(Gathering.participants).selectinload(Participant.user)
    )
    if date_from is not None:
        stmt = stmt.where(Gathering.event_date >= date_from)
    if date_to is not None:
        stmt = stmt.where(Gathering.event_date <= date_to)
    stmt = stmt.order_by(Gathering.event_date, Gathering.start_time)

    return [_to_read(g) for g in db.scalars(stmt).all()]


@router.get("/{gathering_id}", response_model=GatheringDetail)
def get_gathering(
    gathering_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """모임 상세 (참여자 명단 포함)."""
    g = _load_gathering(db, gathering_id)
    detail = GatheringDetail.model_validate(g)
    detail.attendance = _summary(g)
    detail.participants = [ParticipantRead.model_validate(p) for p in g.participants]
    return detail


@router.patch("/{gathering_id}", response_model=GatheringRead)
def update_gathering(
    gathering_id: int,
    payload: GatheringUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    g = _load_gathering(db, gathering_id)
    _require_organizer(g, current_user)
    data = payload.model_dump(exclude_unset=True)
    _normalize_courts(data)
    for field_name, value in data.items():
        setattr(g, field_name, value)
    db.commit()
    return _to_read(_load_gathering(db, gathering_id))


@router.delete("/{gathering_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_gathering(
    gathering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    g = _load_gathering(db, gathering_id)
    _require_organizer(g, current_user)
    db.delete(g)
    db.commit()


# --- 참석 투표 --------------------------------------------------------------
@router.get("/{gathering_id}/participants", response_model=list[ParticipantRead])
def list_participants(
    gathering_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    g = _load_gathering(db, gathering_id)
    return [ParticipantRead.model_validate(p) for p in g.participants]


@router.put("/{gathering_id}/attendance", response_model=ParticipantRead)
def vote_attendance(
    gathering_id: int,
    payload: ParticipantVote,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """현재 회원의 참석/불참/미정 투표 (없으면 생성, 있으면 갱신)."""
    gathering = db.get(Gathering, gathering_id)
    if not gathering:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="모임을 찾을 수 없습니다.")

    participant = db.scalar(
        select(Participant).where(
            Participant.gathering_id == gathering_id,
            Participant.user_id == current_user.id,
        )
    )

    # 마감 제한: 모임 3일 전부터는 일반 회원이 불참/미정으로 바꿀 수 없음(관리자만 가능)
    if (
        payload.status in (AttendanceStatus.ABSENT, AttendanceStatus.MAYBE)
        and not current_user.is_admin
        and _attendance_locked(gathering)
    ):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            detail=f"모임 시작 {ATTENDANCE_LOCK_DAYS}일 전부터는 불참/미정으로 변경할 수 없습니다. 관리자에게 문의하세요.",
        )

    # 정원 제한: '참석'으로 바꾸려는데 이미 정원이 찼고, 본인이 아직 참석이 아니면 차단
    if payload.status == AttendanceStatus.ATTENDING and gathering.max_participants is not None:
        already_attending = (
            participant is not None and participant.status == AttendanceStatus.ATTENDING
        )
        if not already_attending:
            current_attending = db.scalar(
                select(func.count())
                .select_from(Participant)
                .where(
                    Participant.gathering_id == gathering_id,
                    Participant.status == AttendanceStatus.ATTENDING,
                )
            )
            if current_attending >= gathering.max_participants:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    detail=f"참석 정원({gathering.max_participants}명)이 가득 찼습니다.",
                )

    if participant is None:
        participant = Participant(
            gathering_id=gathering_id, user_id=current_user.id, status=payload.status
        )
        db.add(participant)
    else:
        participant.status = payload.status

    db.commit()
    db.refresh(participant)
    return ParticipantRead.model_validate(participant)


@router.delete("/{gathering_id}/attendance", status_code=status.HTTP_204_NO_CONTENT)
def cancel_attendance(
    gathering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """투표 취소 (참여자 명단에서 제거)."""
    gathering = db.get(Gathering, gathering_id)
    # 마감 제한: 취소(=사실상 불참)도 3일 전부터는 일반 회원 불가(관리자만 가능)
    if (
        gathering is not None
        and not current_user.is_admin
        and _attendance_locked(gathering)
    ):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            detail=f"모임 시작 {ATTENDANCE_LOCK_DAYS}일 전부터는 참석 취소를 할 수 없습니다. 관리자에게 문의하세요.",
        )

    participant = db.scalar(
        select(Participant).where(
            Participant.gathering_id == gathering_id,
            Participant.user_id == current_user.id,
        )
    )
    if participant:
        db.delete(participant)
        db.commit()
